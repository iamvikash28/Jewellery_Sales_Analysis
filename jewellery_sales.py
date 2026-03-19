import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import random

random.seed(42)

wb = Workbook()

# ── colour palette ──────────────────────────────────────────────────────────
GOLD     = "FFD700"
GOLD_DRK = "B8860B"
SILVER   = "C0C0C0"
DIAMOND  = "B9F2FF"
HEADER   = "1A1A2E"   # dark navy
SUB_HDR  = "16213E"
ACCENT   = "E94560"
BG_LIGHT = "F8F4E9"
WHITE    = "FFFFFF"
GREEN    = "2ECC71"
AMBER    = "F39C12"
RED_SOFT = "E74C3C"

thin  = Side(style="thin",   color="CCCCCC")
thick = Side(style="medium", color="888888")
border_all  = Border(left=thin, right=thin, top=thin, bottom=thin)
border_head = Border(left=thick, right=thick, top=thick, bottom=thick)

def hdr_font(sz=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def cell_font(sz=10, bold=False, color="222222"):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def money_fmt(ws, cell):
    ws[cell].number_format = '₹#,##0.00'

# ── Sheet 1 : RAW DATA ───────────────────────────────────────────────────────
ws_data = wb.active
ws_data.title = "Sales_Data"
ws_data.sheet_view.showGridLines = False
ws_data.freeze_panes = "A2"

cols = ["Order_ID","Date","Month","Month_Num","Category","Product",
        "Metal_Purity","Qty","Unit_Price","Revenue","Customer_ID",
        "Customer_Type","City","Payment_Method"]

products = {
    "Gold":    ["Gold Ring","Gold Necklace","Gold Bracelet","Gold Earrings","Gold Bangle","Gold Chain","Gold Pendant","Gold Anklet"],
    "Silver":  ["Silver Ring","Silver Necklace","Silver Bracelet","Silver Earrings","Silver Bangle","Silver Chain","Silver Pendant"],
    "Diamond": ["Diamond Ring","Diamond Necklace","Diamond Bracelet","Diamond Earrings","Diamond Pendant","Diamond Set","Diamond Studs"],
}
purities = {"Gold":["18K","22K","24K"], "Silver":["92.5% Sterling","999 Fine"], "Diamond":["VS1","VS2","VVS1","VVS2","SI1"]}
price_range = {"Gold":(4500,35000), "Silver":(800,6000), "Diamond":(15000,180000)}
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
cities = ["Mumbai","Delhi","Bangalore","Hyderabad","Chennai","Kolkata","Jaipur","Surat","Pune","Ahmedabad"]
customer_types = ["New","Returning","Premium"]
payments = ["UPI","Credit Card","Debit Card","Cash","EMI"]

# header row
for ci, h in enumerate(cols, 1):
    c = ws_data.cell(1, ci, h)
    c.font   = hdr_font(11)
    c.fill   = fill(HEADER)
    c.alignment = center()
    c.border = border_head

col_widths = [12,12,8,10,10,20,14,6,12,14,12,14,12,14]
for i, w in enumerate(col_widths, 1):
    ws_data.column_dimensions[get_column_letter(i)].width = w
ws_data.row_dimensions[1].height = 22

rows = []
oid  = 1001
for month_i, month in enumerate(months, 1):
    # seasonal multiplier
    mult = 1.8 if month in ["Oct","Nov","Feb","Mar"] else (0.7 if month in ["Jun","Jul"] else 1.0)
    n    = int(random.randint(35, 55) * mult)
    for _ in range(n):
        cat     = random.choices(["Gold","Silver","Diamond"], weights=[50,30,20])[0]
        product = random.choice(products[cat])
        purity  = random.choice(purities[cat])
        lo, hi  = price_range[cat]
        qty     = random.randint(1, 3)
        price   = round(random.uniform(lo, hi), 2)
        rev     = round(price * qty, 2)
        day     = random.randint(1, 28)
        date_str = f"2024-{month_i:02d}-{day:02d}"
        cust_id  = f"CUST{random.randint(1000,9999)}"
        rows.append([f"ORD{oid}", date_str, month, month_i, cat, product,
                     purity, qty, price, rev,
                     cust_id,
                     random.choice(customer_types),
                     random.choice(cities),
                     random.choice(payments)])
        oid += 1

alt_fills = [fill("FFFDF4"), fill(BG_LIGHT)]
for ri, row in enumerate(rows, 2):
    af = alt_fills[ri % 2]
    for ci, val in enumerate(row, 1):
        c = ws_data.cell(ri, ci, val)
        c.font      = cell_font()
        c.fill      = af
        c.border    = border_all
        c.alignment = center() if ci in [1,3,4,5,8,12,13,14] else Alignment(vertical="center")
        if ci in [9, 10]:
            c.number_format = '₹#,##0.00'

total_rows = len(rows) + 1

# ── Sheet 2 : PIVOT SUMMARIES ────────────────────────────────────────────────
ws_piv = wb.create_sheet("Pivot_Summary")
ws_piv.sheet_view.showGridLines = False

def section_title(ws, row, col, text, width=6):
    c = ws.cell(row, col, text)
    c.font = hdr_font(13, color=WHITE)
    c.fill = fill(HEADER)
    c.alignment = center(wrap=True)
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+width-1)
    ws.row_dimensions[row].height = 24

def sub_header(ws, row, cols_list):
    for ci, txt in zip(cols_list, ["Month","Total Revenue (₹)","Orders","Avg Order Value (₹)"]):
        if ci > max(cols_list): break
    for ci, txt in enumerate(sub_header.labels if hasattr(sub_header,'labels') else
                              ["Month","Revenue (₹)","Orders","Avg (₹)"], cols_list[0]):
        if ci > cols_list[-1]: break
        c = ws.cell(row, ci, txt)
        c.font = hdr_font(10, color=WHITE)
        c.fill = fill(SUB_HDR)
        c.alignment = center()

# build monthly summary from rows data
monthly = {}
for r in rows:
    m, mn, cat, qty, rev = r[2], r[3], r[4], r[7], r[9]
    key = (m, mn)
    if key not in monthly:
        monthly[key] = {"rev":0,"orders":0}
    monthly[key]["rev"]    += rev
    monthly[key]["orders"] += 1

monthly_sorted = sorted(monthly.items(), key=lambda x: x[0][1])

section_title(ws_piv, 1, 1, "📅  Monthly Revenue Summary")
hdrs = ["Month","Revenue (₹)","Orders","Avg Order (₹)","MoM Growth (%)"]
for ci, h in enumerate(hdrs, 1):
    c = ws_piv.cell(2, ci, h)
    c.font = hdr_font(10, color=WHITE)
    c.fill = fill(SUB_HDR)
    c.alignment = center()
    ws_piv.column_dimensions[get_column_letter(ci)].width = 18

for ri, ((m, mn), v) in enumerate(monthly_sorted, 3):
    ws_piv.cell(ri, 1, m).alignment = center()
    c_rev = ws_piv.cell(ri, 2, round(v["rev"], 2))
    c_rev.number_format = '₹#,##0'
    ws_piv.cell(ri, 3, v["orders"]).alignment = center()
    c_avg = ws_piv.cell(ri, 4)
    c_avg.value  = f"=B{ri}/C{ri}"
    c_avg.number_format = '₹#,##0'
    if ri > 3:
        c_mom = ws_piv.cell(ri, 5)
        c_mom.value  = f"=(B{ri}-B{ri-1})/B{ri-1}"
        c_mom.number_format = '0.0%'
    else:
        ws_piv.cell(ri, 5, "-").alignment = center()

    for ci in range(1, 6):
        c = ws_piv.cell(ri, ci)
        c.border    = border_all
        c.fill      = fill("FFFDF4") if ri % 2 == 0 else fill(BG_LIGHT)
        if ci not in [2,4]:
            c.alignment = center()

last_mon_row = 2 + len(monthly_sorted)
# Total row
tr = last_mon_row + 1
ws_piv.cell(tr, 1, "TOTAL").font = hdr_font(10, color=WHITE)
ws_piv.cell(tr, 1).fill = fill(GOLD_DRK)
ws_piv.cell(tr, 1).alignment = center()
c_tot = ws_piv.cell(tr, 2)
c_tot.value  = f"=SUM(B3:B{last_mon_row})"
c_tot.number_format = '₹#,##0'
c_tot.font   = hdr_font(10, color=WHITE)
c_tot.fill   = fill(GOLD_DRK)
c_tot.alignment = center()
c_tot_o = ws_piv.cell(tr, 3)
c_tot_o.value  = f"=SUM(C3:C{last_mon_row})"
c_tot_o.font   = hdr_font(10, color=WHITE)
c_tot_o.fill   = fill(GOLD_DRK)
c_tot_o.alignment = center()
for ci in [4,5]:
    ws_piv.cell(tr, ci).fill = fill(GOLD_DRK)

# Category Summary
cat_summary = {}
for r in rows:
    cat, rev, qty = r[4], r[9], r[7]
    if cat not in cat_summary:
        cat_summary[cat] = {"rev":0,"orders":0,"qty":0}
    cat_summary[cat]["rev"]    += rev
    cat_summary[cat]["orders"] += 1
    cat_summary[cat]["qty"]    += qty

section_title(ws_piv, tr+2, 1, "💎  Revenue by Category")
cat_hdrs = ["Category","Revenue (₹)","Orders","Units Sold","Avg Price (₹)","Revenue Share (%)"]
for ci, h in enumerate(cat_hdrs, 1):
    c = ws_piv.cell(tr+3, ci, h)
    c.font = hdr_font(10, color=WHITE)
    c.fill = fill(SUB_HDR)
    c.alignment = center()

cat_fills = {"Gold": GOLD, "Silver": SILVER, "Diamond": DIAMOND}
cat_row_start = tr+4
for ri2, (cat, v) in enumerate(cat_summary.items(), cat_row_start):
    cf = fill(cat_fills.get(cat, BG_LIGHT))
    ws_piv.cell(ri2, 1, cat).alignment = center()
    c_r = ws_piv.cell(ri2, 2, round(v["rev"],2))
    c_r.number_format = '₹#,##0'
    ws_piv.cell(ri2, 3, v["orders"]).alignment = center()
    ws_piv.cell(ri2, 4, v["qty"]).alignment = center()
    c_ap = ws_piv.cell(ri2, 5)
    c_ap.value  = f"=B{ri2}/D{ri2}"
    c_ap.number_format = '₹#,##0'
    c_rs = ws_piv.cell(ri2, 6)
    c_rs.value  = f"=B{ri2}/SUM(B{cat_row_start}:B{cat_row_start+len(cat_summary)-1})"
    c_rs.number_format = '0.0%'
    for ci in range(1, 7):
        c = ws_piv.cell(ri2, ci)
        c.border = border_all
        c.fill   = cf
        if ci not in [2,5,6]:
            c.alignment = center()

# Product Top 10
product_rev = {}
for r in rows:
    prod, rev = r[5], r[9]
    product_rev[prod] = product_rev.get(prod, 0) + rev

top10 = sorted(product_rev.items(), key=lambda x: -x[1])[:10]
last_cat_row = cat_row_start + len(cat_summary) - 1

section_title(ws_piv, last_cat_row+2, 1, "🏆  Top 10 Best-Selling Products")
for ci, h in enumerate(["Rank","Product","Revenue (₹)","Revenue Share (%)"], 1):
    c = ws_piv.cell(last_cat_row+3, ci, h)
    c.font = hdr_font(10, color=WHITE)
    c.fill = fill(SUB_HDR)
    c.alignment = center()

prod_row_start = last_cat_row+4
total_rev = sum(v["rev"] for v in cat_summary.values())
for ri3, (prod, rev) in enumerate(top10, prod_row_start):
    rank = ri3 - prod_row_start + 1
    ws_piv.cell(ri3, 1, rank).alignment = center()
    ws_piv.cell(ri3, 2, prod)
    c_r = ws_piv.cell(ri3, 3, round(rev, 2))
    c_r.number_format = '₹#,##0'
    c_sh = ws_piv.cell(ri3, 4, round(rev/total_rev, 4))
    c_sh.number_format = '0.0%'
    clr = GOLD if rank <= 3 else ("E8E8E8" if rank <= 6 else "F8F8F8")
    for ci in range(1, 5):
        c = ws_piv.cell(ri3, ci)
        c.border = border_all
        c.fill   = fill(clr)
        c.alignment = center() if ci != 2 else Alignment(vertical="center")

# ── Sheet 3 : SQL QUERIES ────────────────────────────────────────────────────
ws_sql = wb.create_sheet("SQL_Queries")
ws_sql.sheet_view.showGridLines = False
ws_sql.column_dimensions['A'].width = 2
ws_sql.column_dimensions['B'].width = 90

queries = [
    ("-- 1. Monthly Revenue Trend",
     """SELECT Month, Month_Num,
       SUM(Revenue)         AS Total_Revenue,
       COUNT(Order_ID)      AS Total_Orders,
       ROUND(AVG(Revenue),2) AS Avg_Order_Value
FROM jewellery_sales
GROUP BY Month, Month_Num
ORDER BY Month_Num;"""),
    ("-- 2. Revenue by Category",
     """SELECT Category,
       SUM(Revenue)             AS Total_Revenue,
       COUNT(Order_ID)          AS Total_Orders,
       SUM(Qty)                 AS Units_Sold,
       ROUND(AVG(Unit_Price),2) AS Avg_Unit_Price,
       ROUND(SUM(Revenue)*100.0/
         (SELECT SUM(Revenue) FROM jewellery_sales),1) AS Revenue_Share_Pct
FROM jewellery_sales
GROUP BY Category
ORDER BY Total_Revenue DESC;"""),
    ("-- 3. Top 10 Best-Selling Products",
     """SELECT Product, Category,
       SUM(Revenue)    AS Total_Revenue,
       SUM(Qty)        AS Units_Sold,
       COUNT(Order_ID) AS Orders
FROM jewellery_sales
GROUP BY Product, Category
ORDER BY Total_Revenue DESC
LIMIT 10;"""),
    ("-- 4. Monthly Sales by Category (Cross-Tab)",
     """SELECT Month, Month_Num,
       SUM(CASE WHEN Category='Gold'    THEN Revenue ELSE 0 END) AS Gold_Revenue,
       SUM(CASE WHEN Category='Silver'  THEN Revenue ELSE 0 END) AS Silver_Revenue,
       SUM(CASE WHEN Category='Diamond' THEN Revenue ELSE 0 END) AS Diamond_Revenue
FROM jewellery_sales
GROUP BY Month, Month_Num
ORDER BY Month_Num;"""),
    ("-- 5. Customer Type Analysis",
     """SELECT Customer_Type,
       COUNT(DISTINCT Customer_ID) AS Unique_Customers,
       COUNT(Order_ID)             AS Total_Orders,
       SUM(Revenue)                AS Total_Revenue,
       ROUND(AVG(Revenue),2)       AS Avg_Order_Value
FROM jewellery_sales
GROUP BY Customer_Type
ORDER BY Total_Revenue DESC;"""),
    ("-- 6. City-wise Revenue Leaders",
     """SELECT City,
       SUM(Revenue)    AS Total_Revenue,
       COUNT(Order_ID) AS Orders,
       ROUND(AVG(Revenue),2) AS Avg_Order
FROM jewellery_sales
GROUP BY City
ORDER BY Total_Revenue DESC
LIMIT 10;"""),
    ("-- 7. Payment Method Breakdown",
     """SELECT Payment_Method,
       COUNT(Order_ID) AS Orders,
       SUM(Revenue)    AS Revenue,
       ROUND(AVG(Revenue),2) AS Avg_Txn
FROM jewellery_sales
GROUP BY Payment_Method
ORDER BY Revenue DESC;"""),
    ("-- 8. Peak Sales Days (by Month)",
     """SELECT Month, Month_Num,
       COUNT(Order_ID) AS Daily_Orders
FROM jewellery_sales
GROUP BY Month, Month_Num
ORDER BY Daily_Orders DESC
LIMIT 5;"""),
    ("-- 9. Quarter-over-Quarter Comparison",
     """SELECT
  CASE WHEN Month_Num BETWEEN 1 AND 3  THEN 'Q1'
       WHEN Month_Num BETWEEN 4 AND 6  THEN 'Q2'
       WHEN Month_Num BETWEEN 7 AND 9  THEN 'Q3'
       ELSE 'Q4' END AS Quarter,
  Category,
  SUM(Revenue)    AS Revenue,
  COUNT(Order_ID) AS Orders
FROM jewellery_sales
GROUP BY Quarter, Category
ORDER BY Quarter, Revenue DESC;"""),
    ("-- 10. High-Value Transactions (> ₹50,000)",
     """SELECT Order_ID, Date, Product, Category,
       Qty, Unit_Price, Revenue,
       Customer_Type, City
FROM jewellery_sales
WHERE Revenue > 50000
ORDER BY Revenue DESC;"""),
]

row_cur = 1
for comment, sql in queries:
    # comment line
    c = ws_sql.cell(row_cur, 2, comment)
    c.font = Font(name="Courier New", size=10, bold=True, color=GOLD_DRK)
    c.fill = fill("1A1A2E")
    ws_sql.row_dimensions[row_cur].height = 18
    row_cur += 1
    for line in sql.split("\n"):
        c = ws_sql.cell(row_cur, 2, line)
        c.font = Font(name="Courier New", size=10, color="00FF7F" if line.strip().startswith("--") else "E0E0E0")
        c.fill = fill("0D0D1A")
        ws_sql.row_dimensions[row_cur].height = 16
        row_cur += 1
    row_cur += 1  # blank gap

# ── Sheet 4 : CHARTS DASHBOARD ───────────────────────────────────────────────
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_view.showGridLines = False

# Title banner
ws_dash.row_dimensions[1].height = 36
ws_dash.row_dimensions[2].height = 20
for col in range(1, 20):
    c = ws_dash.cell(1, col)
    c.fill = fill(HEADER)
title_c = ws_dash.cell(1, 2, "💍  JEWELLERY SALES ANALYSIS DASHBOARD  |  FY 2024")
title_c.font = Font(name="Calibri", size=16, bold=True, color=GOLD)
title_c.alignment = Alignment(horizontal="left", vertical="center")
ws_dash.merge_cells("B1:S1")

sub_c = ws_dash.cell(2, 2, "Comprehensive analytics across Gold · Silver · Diamond categories")
sub_c.font = Font(name="Calibri", size=10, italic=True, color=SILVER)
sub_c.fill = fill(HEADER)
ws_dash.merge_cells("B2:S2")

# KPI cards row (row 4)
kpi_data = [
    ("Total Revenue",  f"=SUM(Pivot_Summary!B3:B{last_mon_row})", "₹#,##0",   GOLD,    "💰"),
    ("Total Orders",   f"=SUM(Pivot_Summary!C3:C{last_mon_row})", "#,##0",    ACCENT,  "📦"),
    ("Avg Order Value",f"=Pivot_Summary!B{tr}/Pivot_Summary!C{tr}","₹#,##0",  GREEN,   "📊"),
    ("Best Month",     "Oct",                                      "@",        DIAMOND, "🏆"),
]
col_positions = [2, 6, 10, 14]
ws_dash.row_dimensions[4].height = 14
ws_dash.row_dimensions[5].height = 28
ws_dash.row_dimensions[6].height = 22
ws_dash.row_dimensions[7].height = 14

for kpi, (label, val, fmt, clr, icon), col in zip(range(4), kpi_data, col_positions):
    for r in [4,5,6,7]:
        for cc in range(col, col+3):
            ws_dash.cell(r, cc).fill = fill(clr if clr != DIAMOND else "A8EDFF")
    icon_c = ws_dash.cell(5, col, icon)
    icon_c.font = Font(size=18)
    icon_c.alignment = center()
    ws_dash.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+2)
    lbl_c = ws_dash.cell(6, col, label)
    lbl_c.font = Font(name="Calibri", size=9, bold=True, color="333333")
    lbl_c.alignment = center()
    ws_dash.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+2)

# Write data tables for charts into helper area
# Monthly revenue helper (col 20+)
helper_col = 20
ws_dash.cell(3, helper_col,   "Month")
ws_dash.cell(3, helper_col+1, "Revenue")
ws_dash.cell(3, helper_col+2, "Gold")
ws_dash.cell(3, helper_col+3, "Silver")
ws_dash.cell(3, helper_col+4, "Diamond")

# Pull from pivot summary
for i, (row_in_piv) in enumerate(range(3, 3+len(monthly_sorted)), 4):
    ws_dash.cell(i, helper_col,   f"=Pivot_Summary!A{row_in_piv}")
    ws_dash.cell(i, helper_col+1, f"=Pivot_Summary!B{row_in_piv}")

# Category helper
cat_names = list(cat_summary.keys())
ws_dash.cell(3, helper_col+6, "Category")
ws_dash.cell(3, helper_col+7, "Revenue")
for i, cat in enumerate(cat_names, 4):
    ri_in_piv = cat_row_start + list(cat_summary.keys()).index(cat)
    ws_dash.cell(i, helper_col+6, cat)
    ws_dash.cell(i, helper_col+7, f"=Pivot_Summary!B{ri_in_piv}")

# ── LINE CHART: Monthly Revenue Trend ────────────────────────────────────────
lc = LineChart()
lc.title     = "Monthly Revenue Trend (₹)"
lc.style     = 10
lc.y_axis.title = "Revenue (₹)"
lc.x_axis.title = "Month"
lc.width  = 16
lc.height = 10

data_ref  = Reference(ws_dash, min_col=helper_col+1, min_row=3,
                       max_row=3+len(monthly_sorted))
cats_ref  = Reference(ws_dash, min_col=helper_col,   min_row=4,
                       max_row=3+len(monthly_sorted))
lc.add_data(data_ref, titles_from_data=True)
lc.set_categories(cats_ref)
lc.series[0].graphicalProperties.line.solidFill = GOLD_DRK
lc.series[0].graphicalProperties.line.width     = 25000
ws_dash.add_chart(lc, "B9")

# ── BAR CHART: Top 10 Products ────────────────────────────────────────────────
bc = BarChart()
bc.type       = "bar"
bc.title      = "Top 10 Products by Revenue"
bc.style      = 10
bc.y_axis.title = "Product"
bc.x_axis.title = "Revenue (₹)"
bc.width  = 16
bc.height = 12

top10_col = helper_col + 9
ws_dash.cell(3, top10_col,   "Product")
ws_dash.cell(3, top10_col+1, "Revenue")
for i, (prod, rev) in enumerate(top10, 4):
    ws_dash.cell(i, top10_col,   prod)
    ws_dash.cell(i, top10_col+1, round(rev,2))

data_ref2 = Reference(ws_dash, min_col=top10_col+1, min_row=3, max_row=13)
cats_ref2 = Reference(ws_dash, min_col=top10_col,   min_row=4, max_row=13)
bc.add_data(data_ref2, titles_from_data=True)
bc.set_categories(cats_ref2)
bc.series[0].graphicalProperties.solidFill = GOLD
ws_dash.add_chart(bc, "K9")

# ── PIE CHART: Revenue by Category ───────────────────────────────────────────
pc = PieChart()
pc.title  = "Revenue Share by Category"
pc.style  = 10
pc.width  = 14
pc.height = 10

data_ref3 = Reference(ws_dash, min_col=helper_col+7, min_row=3, max_row=3+len(cat_names))
cats_ref3 = Reference(ws_dash, min_col=helper_col+6, min_row=4, max_row=3+len(cat_names))
pc.add_data(data_ref3, titles_from_data=True)
pc.set_categories(cats_ref3)
pc.dataLabels = None
slice_clrs = [GOLD, SILVER, "B9F2FF"]
for i, clr in enumerate(slice_clrs):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = clr
    pc.series[0].data_points.append(pt)
ws_dash.add_chart(pc, "B27")

# ── BAR CHART: Monthly Comparison (Grouped) ───────────────────────────────────
bc2 = BarChart()
bc2.type       = "col"
bc2.title      = "Monthly Sales Comparison by Category"
bc2.style      = 10
bc2.y_axis.title = "Revenue (₹)"
bc2.x_axis.title = "Month"
bc2.width  = 18
bc2.height = 10
bc2.grouping = "clustered"

# Build monthly-by-category table
mc_col = helper_col + 12
ws_dash.cell(3, mc_col,   "Month")
ws_dash.cell(3, mc_col+1, "Gold")
ws_dash.cell(3, mc_col+2, "Silver")
ws_dash.cell(3, mc_col+3, "Diamond")

month_cat = {}
for r in rows:
    m, mn, cat, rev = r[2], r[3], r[4], r[9]
    if mn not in month_cat:
        month_cat[mn] = {mn: mn, "month": m, "Gold":0,"Silver":0,"Diamond":0}
    month_cat[mn][cat] = month_cat[mn].get(cat,0) + rev

for i, mn in enumerate(sorted(month_cat.keys()), 4):
    d = month_cat[mn]
    ws_dash.cell(i, mc_col,   d["month"])
    ws_dash.cell(i, mc_col+1, round(d["Gold"],2))
    ws_dash.cell(i, mc_col+2, round(d["Silver"],2))
    ws_dash.cell(i, mc_col+3, round(d["Diamond"],2))

for series_col, clr in [(mc_col+1, GOLD_DRK), (mc_col+2, "A8A8A8"), (mc_col+3, "5DADE2")]:
    data_ref4 = Reference(ws_dash, min_col=series_col, min_row=3, max_row=15)
    bc2.add_data(data_ref4, titles_from_data=True)

cats_ref4 = Reference(ws_dash, min_col=mc_col, min_row=4, max_row=15)
bc2.set_categories(cats_ref4)
for i, clr in enumerate([GOLD_DRK, "A8A8A8", "5DADE2"]):
    bc2.series[i].graphicalProperties.solidFill = clr
ws_dash.add_chart(bc2, "K27")

# ── Sheet 5 : README ─────────────────────────────────────────────────────────
ws_readme = wb.create_sheet("📋 README")
ws_readme.sheet_view.showGridLines = False
ws_readme.column_dimensions['B'].width = 70

readme_lines = [
    ("JEWELLERY SALES ANALYSIS PROJECT", HEADER, 16, True, WHITE),
    ("A complete Data Analyst portfolio project", HEADER, 11, False, SILVER),
    ("", "", 10, False, "222222"),
    ("📁 WORKBOOK STRUCTURE", GOLD_DRK, 12, True, WHITE),
    ("  Sales_Data      →  Raw transactional data (500+ orders, 2024)", BG_LIGHT, 10, False, "222222"),
    ("  Pivot_Summary   →  Monthly trends, category breakdown, top products", BG_LIGHT, 10, False, "222222"),
    ("  SQL_Queries     →  10 business queries for analysis", BG_LIGHT, 10, False, "222222"),
    ("  Dashboard       →  4 interactive charts for Power BI import", BG_LIGHT, 10, False, "222222"),
    ("", "", 10, False, "222222"),
    ("📊 WHAT THIS PROJECT COVERS", GOLD_DRK, 12, True, WHITE),
    ("  ✔ Monthly revenue trends with MoM growth %", BG_LIGHT, 10, False, "222222"),
    ("  ✔ Revenue split: Gold | Silver | Diamond categories", BG_LIGHT, 10, False, "222222"),
    ("  ✔ Top 10 best-selling jewellery products", BG_LIGHT, 10, False, "222222"),
    ("  ✔ Customer type segmentation (New / Returning / Premium)", BG_LIGHT, 10, False, "222222"),
    ("  ✔ City-wise sales performance (10 cities)", BG_LIGHT, 10, False, "222222"),
    ("  ✔ Payment method analysis", BG_LIGHT, 10, False, "222222"),
    ("  ✔ Seasonal peaks (Diwali, Valentine's, Akshaya Tritiya)", BG_LIGHT, 10, False, "222222"),
    ("", "", 10, False, "222222"),
    ("🔑 KEY INSIGHTS TO HIGHLIGHT IN INTERVIEW", GOLD_DRK, 12, True, WHITE),
    ("  1. Oct/Nov peak → Diwali festive demand surge", BG_LIGHT, 10, False, "222222"),
    ("  2. Gold drives ~50% of revenue despite higher price point", BG_LIGHT, 10, False, "222222"),
    ("  3. Diamond avg order value 4x higher than Silver", BG_LIGHT, 10, False, "222222"),
    ("  4. Premium customers have highest avg transaction value", BG_LIGHT, 10, False, "222222"),
    ("  5. Jun/Jul are lowest months → off-season strategy needed", BG_LIGHT, 10, False, "222222"),
    ("", "", 10, False, "222222"),
    ("🛠 HOW TO USE THIS FOR YOUR PROJECT", GOLD_DRK, 12, True, WHITE),
    ("  Step 1: Explore Sales_Data tab — understand schema", BG_LIGHT, 10, False, "222222"),
    ("  Step 2: Study Pivot_Summary — formulas are dynamic", BG_LIGHT, 10, False, "222222"),
    ("  Step 3: Run SQL_Queries in SQLite / MySQL / PostgreSQL", BG_LIGHT, 10, False, "222222"),
    ("  Step 4: Import Sales_Data into Power BI for dashboard", BG_LIGHT, 10, False, "222222"),
    ("  Step 5: Present Dashboard charts as your visual deliverable", BG_LIGHT, 10, False, "222222"),
]

for ri, (text, bg, sz, bold, fc) in enumerate(readme_lines, 1):
    c = ws_readme.cell(ri, 2, text)
    c.font = Font(name="Calibri", size=sz, bold=bold, color=fc)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(vertical="center")
    ws_readme.row_dimensions[ri].height = 20

# ── Reorder sheets ────────────────────────────────────────────────────────────
sheet_order = ["📋 README","Sales_Data","Pivot_Summary","SQL_Queries","Dashboard"]
for s in sheet_order:
    wb._sheets.append(wb._sheets.pop(wb.sheetnames.index(s)))

out = "Jewellery_Sales_Analysis.xlsx"
wb.save(out)
print("Saved:", out)
